import csv
import io
import re
import openpyxl
from django.db.models import Q
from .models import Contact, Category, ImportLog

MAX_IMPORT_ROWS = 5000


def _get_or_create_category(user, name):
    name = (name or '').strip()
    if not name:
        return None
    default = Category.objects.filter(user=None, name__iexact=name).first()
    if default:
        return default
    user_cat = Category.objects.filter(user=user, name__iexact=name).first()
    if user_cat:
        return user_cat
    return Category.objects.create(user=user, name=name)


def _is_valid_email(email):
    if not email:
        return True
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email))


def _is_duplicate(user, phone, email, exclude_id=None):
    """
    Used by the preview step to count how many rows already exist.
    Exact phone OR exact email match counts as duplicate.
    """
    qs = Contact.objects.filter(user=user)
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    if phone and qs.filter(phone_number=phone).exists():
        return True
    if email and qs.filter(email=email).exists():
        return True
    return False


def _is_exact_duplicate(user, name, phone, email):
    """
    Stricter check used at commit time.
    Only skips a row if the name AND (phone or email) both match an existing contact.
    This allows 'Mum, 08012345678' and 'Mama, 08012345678' to both be imported
    so the user can review and merge them in Settings.
    """
    qs = Contact.objects.filter(user=user)
    if phone:
        match = qs.filter(phone_number=phone, full_name__iexact=name).first()
        if match:
            return True
    if email:
        match = qs.filter(email__iexact=email, full_name__iexact=name).first()
        if match:
            return True
    return False


def _normalise_row(name, phone, email, category):
    name     = (name     or '').strip()
    phone    = (phone    or '').strip()
    email    = (email    or '').strip()
    category = (category or '').strip()
    if not name or not phone:
        return None
    return {'name': name, 'phone': phone, 'email': email, 'category': category}


def parse_file_for_preview(file, filename):
    filename = filename.lower()
    if filename.endswith('.csv'):
        return _parse_csv(file)
    elif filename.endswith('.txt'):
        return _parse_txt(file)
    elif filename.endswith('.xlsx'):
        return _parse_xlsx(file)
    else:
        return {'rows': [], 'failed': 0, 'error': 'Unsupported file type. Use CSV, TXT, or XLSX.', 'format': None}


def _parse_csv(file):
    rows   = []
    failed = 0
    try:
        reader = csv.DictReader(io.TextIOWrapper(file, encoding='utf-8-sig', newline=''))
    except Exception:
        return {'rows': [], 'failed': 0, 'error': 'Could not read CSV file.', 'format': 'csv'}

    for index, raw in enumerate(reader, start=1):
        if index > MAX_IMPORT_ROWS:
            return {
                'rows': [],
                'failed': 0,
                'error': f'Import limit exceeded. Please upload at most {MAX_IMPORT_ROWS:,} rows at a time.',
                'format': 'csv',
            }
        row = _normalise_row(
            raw.get('name') or raw.get('Name'),
            raw.get('phone') or raw.get('Phone'),
            raw.get('email') or raw.get('Email'),
            raw.get('category') or raw.get('Category'),
        )
        if row is None or not _is_valid_email(row['email']):
            failed += 1
        else:
            rows.append(row)

    return {'rows': rows, 'failed': failed, 'error': None, 'format': 'csv'}


def _parse_txt(file):
    rows   = []
    failed = 0
    try:
        lines = io.TextIOWrapper(file, encoding='utf-8-sig', newline='')
    except Exception:
        return {'rows': [], 'failed': 0, 'error': 'Could not read TXT file.', 'format': 'txt'}

    for index, line in enumerate(lines, start=1):
        if index > MAX_IMPORT_ROWS:
            return {
                'rows': [],
                'failed': 0,
                'error': f'Import limit exceeded. Please upload at most {MAX_IMPORT_ROWS:,} rows at a time.',
                'format': 'txt',
            }
        line = line.strip()
        if not line:
            continue
        parts = line.split('|')
        if len(parts) < 2:
            failed += 1
            continue
        row = _normalise_row(
            parts[0] if len(parts) > 0 else '',
            parts[1] if len(parts) > 1 else '',
            parts[2] if len(parts) > 2 else '',
            parts[3] if len(parts) > 3 else '',
        )
        if row is None or not _is_valid_email(row['email']):
            failed += 1
        else:
            rows.append(row)

    return {'rows': rows, 'failed': failed, 'error': None, 'format': 'txt'}


def _parse_xlsx(file):
    rows   = []
    failed = 0
    try:
        wb      = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws      = wb.active
        headers = None
        for i, excel_row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip().lower() if h else '' for h in excel_row]
                continue
            if i > MAX_IMPORT_ROWS:
                wb.close()
                return {
                    'rows': [],
                    'failed': 0,
                    'error': f'Import limit exceeded. Please upload at most {MAX_IMPORT_ROWS:,} rows at a time.',
                    'format': 'xlsx',
                }
            if not any(excel_row):
                continue

            def cell(col_name, fallback=''):
                if headers and col_name in headers:
                    val = excel_row[headers.index(col_name)]
                    return str(val).strip() if val is not None else fallback
                return fallback

            row = _normalise_row(
                cell('name'),
                cell('phone'),
                cell('email'),
                cell('category'),
            )
            if row is None or not _is_valid_email(row['email']):
                failed += 1
            else:
                rows.append(row)
        wb.close()
    except Exception as e:
        return {'rows': [], 'failed': 0, 'error': f'Could not read XLSX file: {e}', 'format': 'xlsx'}

    return {'rows': rows, 'failed': failed, 'error': None, 'format': 'xlsx'}


def commit_import(user, rows):
    """
    Commits parsed rows to the database.
    Uses the stricter _is_exact_duplicate check — only skips a row if
    both the name AND phone/email already exist together. This means
    'Mum, 08012345678' and 'Mama, 08012345678' will both be imported,
    creating a flaggable duplicate pair for the Settings merge screen.
    """
    imported = 0
    skipped  = 0
    failed   = 0
    contacts_to_create = []

    phones = {row['phone'] for row in rows if row.get('phone')}
    emails = {row['email'] for row in rows if row.get('email')}
    emails.update(email.lower() for email in list(emails))

    existing_signatures = set()
    existing_contacts = (
        Contact.objects
        .filter(user=user)
        .filter(Q(phone_number__in=phones) | Q(email__in=emails))
        .values_list('full_name', 'phone_number', 'email')
    )
    for name, phone, email in existing_contacts.iterator(chunk_size=1000):
        name_key = (name or '').lower()
        if phone:
            existing_signatures.add((name_key, 'phone', phone))
        if email:
            existing_signatures.add((name_key, 'email', email.lower()))

    categories = {}
    for category in Category.objects.filter(Q(user=None) | Q(user=user)):
        categories[category.name.lower()] = category

    for row in rows:
        try:
            name_key = row['name'].lower()
            email_key = row['email'].lower()
            phone_duplicate = row['phone'] and (name_key, 'phone', row['phone']) in existing_signatures
            email_duplicate = email_key and (name_key, 'email', email_key) in existing_signatures
            if phone_duplicate or email_duplicate:
                skipped += 1
                continue

            cat = None
            category_key = row['category'].strip().lower()
            if category_key:
                cat = categories.get(category_key)
                if cat is None:
                    cat = Category.objects.create(user=user, name=row['category'].strip())
                    categories[category_key] = cat

            contacts_to_create.append(Contact(
                user=user,
                full_name=row['name'],
                phone_number=row['phone'],
                email=row['email'],
                category=cat,
            ))
            if row['phone']:
                existing_signatures.add((name_key, 'phone', row['phone']))
            if email_key:
                existing_signatures.add((name_key, 'email', email_key))
            imported += 1
        except Exception:
            failed += 1

    Contact.objects.bulk_create(contacts_to_create, batch_size=500)

    ImportLog.objects.create(
        user=user,
        imported_count=imported,
        skipped_duplicates=skipped,
        failed_rows=failed,
    )

    return {'imported': imported, 'skipped': skipped, 'failed': failed}


def export_xlsx(user):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title='Contacts')

    headers = ['Name', 'Phone', 'Email', 'Category', 'Date Added']
    ws.append(headers)

    contacts = Contact.objects.filter(user=user).select_related('category').order_by('full_name')
    for c in contacts.iterator(chunk_size=1000):
        ws.append([
            c.full_name,
            c.phone_number,
            c.email,
            c.category.name if c.category else '',
            c.created_at.strftime('%Y-%m-%d'),
        ])

    return wb


def export_csv(user):
    yield ['name', 'phone', 'email', 'category']
    contacts = Contact.objects.filter(user=user).select_related('category').order_by('full_name')
    for c in contacts.iterator(chunk_size=1000):
        yield [
            c.full_name,
            c.phone_number,
            c.email,
            c.category.name if c.category else '',
        ]


def export_txt(user):
    contacts = Contact.objects.filter(user=user).select_related('category').order_by('full_name')
    for c in contacts.iterator(chunk_size=1000):
        yield '|'.join([
            c.full_name,
            c.phone_number,
            c.email,
            c.category.name if c.category else '',
        ]) + '\n'
